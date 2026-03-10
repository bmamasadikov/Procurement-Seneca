from __future__ import annotations

from datetime import datetime
from html import escape
from io import BytesIO
import re
from typing import Dict, Iterable, List, Sequence

import openpyxl
import pandas as pd

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    EXPORT_PDF_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised by runtime config
    EXPORT_PDF_AVAILABLE = False


PROCUREMENT_CONTEXT_COLUMN = "Department"
PROCUREMENT_EXPORT_COLUMNS = [
    "Department",
    "Item",
    "Category",
    "Qty",
    "Unit",
    "Unit Price",
    "Ordered Qty",
    "Total Price",
    "Received Qty",
    "Supplier",
    "PO Number",
]
PROCUREMENT_TABLE_COLUMNS = [
    "Item",
    "Category",
    "Qty",
    "Unit",
    "Unit Price",
    "Ordered Qty",
    "Total Price",
    "Received Qty",
    "Supplier",
    "PO Number",
]

RFP_BASE_EXPORT_COLUMNS = [
    "Item",
    "Qty",
    "Specification",
    "Selected Supplier",
    "Quoted Unit Price",
    "Lead Time",
    "Supplier Notes",
    "RFP Status",
    "RFP Notes",
]

_PRICE_COLUMNS = {"Unit Price", "Total Price", "Quoted Unit Price"}
_QUANTITY_COLUMNS = {"Qty", "Ordered Qty", "Received Qty"}
_WIDE_COLUMN_HINTS = {
    "Item": 2.0,
    "Department": 1.6,
    "Category": 1.3,
    "Supplier": 1.6,
    "Selected Supplier": 1.7,
    "Specification": 2.1,
    "Lead Time": 1.2,
    "Supplier Notes": 1.8,
    "RFP Notes": 1.8,
    "Notes": 1.8,
}


def _safe_float(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _clean_cell(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    return value


def _format_review_status(status: str) -> str:
    labels = {
        "open": "Open",
        "needs_correction": "Needs Correction",
        "corrected": "Corrected",
        "approved": "Approved",
    }
    normalized = str(status or "open").strip()
    return labels.get(normalized, normalized.replace("_", " ").title())


def _normalize_filter_value(value) -> str:
    if isinstance(value, (list, tuple, set)):
        normalized = [str(item).strip() for item in value if str(item).strip()]
        return ", ".join(normalized) or "None"
    text = str(value or "").strip()
    return text or "All"


def sanitize_filename_part(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    text = text.strip("._")
    return text or "export"


def build_export_file_stem(project: Dict, suffix: str) -> str:
    hotel_info = project.get("hotel_info", {})
    hotel_name = hotel_info.get("property_name") or project.get("name") or "hotel"
    project_id = project.get("project_id", "project")
    return f"{sanitize_filename_part(hotel_name)}_{sanitize_filename_part(project_id)}_{suffix}"


def build_procurement_export_df(filtered_items: List[Dict]) -> pd.DataFrame:
    rows = []
    for item in filtered_items:
        item_data = item.get("item_data", {})
        status = item.get("procurement_status", {})
        unit_price = _safe_float(status.get("unit_price", 0))
        ordered_qty = _safe_float(status.get("ordered_qty", 0))
        total_price = _safe_float(status.get("total_price", 0))
        if not total_price and unit_price and ordered_qty:
            total_price = unit_price * ordered_qty
        rows.append(
            {
                "Department": item.get("department", ""),
                "Item": item_data.get("Item", item_data.get("item", "N/A")),
                "Category": item.get("category", ""),
                "Qty": _safe_float(item_data.get("Total Qty", item_data.get("Total_Qty", 0))),
                "Unit": item_data.get("Unit", "pcs"),
                "Unit Price": unit_price,
                "Ordered Qty": ordered_qty,
                "Total Price": total_price,
                "Received Qty": _safe_float(status.get("received_qty", 0)),
                "Supplier": status.get("supplier", ""),
                "PO Number": status.get("po_number", ""),
            }
        )
    return pd.DataFrame(rows, columns=PROCUREMENT_EXPORT_COLUMNS)


def build_rfp_export_df(edited_df: pd.DataFrame) -> pd.DataFrame:
    if edited_df is None or edited_df.empty:
        return pd.DataFrame(columns=RFP_BASE_EXPORT_COLUMNS)
    dynamic_columns = [column for column in edited_df.columns if column not in RFP_BASE_EXPORT_COLUMNS]
    ordered_columns = [column for column in RFP_BASE_EXPORT_COLUMNS if column in edited_df.columns] + dynamic_columns
    normalized = edited_df[ordered_columns].copy()
    return normalized.where(pd.notna(normalized), "")


def build_export_summary_df(kind: str, project: Dict, filters: Dict[str, object], rows_df: pd.DataFrame) -> pd.DataFrame:
    hotel_info = project.get("hotel_info", {})
    rows = [
        {"Metric": "Export Type", "Value": kind},
        {"Metric": "Project ID", "Value": project.get("project_id", "")},
        {"Metric": "Hotel", "Value": hotel_info.get("property_name", "")},
        {"Metric": "Project Name", "Value": hotel_info.get("project_name", "")},
        {"Metric": "Brand", "Value": hotel_info.get("brand", "")},
        {"Metric": "Rooms", "Value": hotel_info.get("total_rooms", 0)},
        {"Metric": "Exported At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Metric": "Row Count", "Value": len(rows_df.index)},
    ]

    for key, value in (filters or {}).items():
        rows.append({"Metric": key, "Value": _normalize_filter_value(value)})

    if PROCUREMENT_CONTEXT_COLUMN in rows_df.columns:
        department_counts = rows_df[PROCUREMENT_CONTEXT_COLUMN].astype(str).fillna("").value_counts()
        for department, count in department_counts.items():
            if department:
                rows.append({"Metric": f"Department: {department}", "Value": int(count)})

    if "RFP Status" in rows_df.columns:
        status_counts = rows_df["RFP Status"].astype(str).fillna("").value_counts()
        for status, count in status_counts.items():
            rows.append({"Metric": f"RFP Status: {status or 'Blank'}", "Value": int(count)})

    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _ordered_unique(values: Iterable[object]) -> List[str]:
    seen = set()
    ordered = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        ordered.append(text)
    return ordered


def _prepare_export_groups(
    rows_df: pd.DataFrame,
    context_column: str | None = None,
    table_columns: Sequence[str] | None = None,
    flat_when_multi_context: bool = False,
):
    display_columns = [column for column in (table_columns or list(rows_df.columns)) if column in rows_df.columns]
    if context_column and context_column in rows_df.columns:
        context_values = _ordered_unique(rows_df[context_column].tolist())
        if len(context_values) == 1:
            single_context = context_values[0]
            display_df = rows_df[display_columns].copy()
            return {
                "context_value": single_context,
                "groups": [(single_context, display_df)],
                "csv_df": display_df,
                "display_columns": display_columns,
            }

        if context_values:
            groups = []
            for context_value in context_values:
                group_df = rows_df[rows_df[context_column].astype(str) == context_value][display_columns].copy()
                groups.append((context_value, group_df))
            csv_columns = ([context_column] if flat_when_multi_context else []) + display_columns
            csv_df = rows_df[csv_columns].copy()
            return {
                "context_value": "",
                "groups": groups,
                "csv_df": csv_df,
                "display_columns": display_columns,
            }

    display_df = rows_df[display_columns].copy() if display_columns else rows_df.copy()
    return {
        "context_value": "",
        "groups": [("", display_df)],
        "csv_df": display_df,
        "display_columns": display_columns or list(display_df.columns),
    }


def generate_csv_bytes(
    rows_df: pd.DataFrame,
    context_column: str | None = None,
    table_columns: Sequence[str] | None = None,
    flat_when_multi_context: bool = False,
) -> bytes:
    prepared = _prepare_export_groups(
        rows_df,
        context_column=context_column,
        table_columns=table_columns,
        flat_when_multi_context=flat_when_multi_context,
    )
    return prepared["csv_df"].to_csv(index=False).encode("utf-8-sig")


def _autofit_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 11), 46)

def _write_report_sheet(workbook, sheet_name: str, title: str, department: str, rows_df: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title=sheet_name[:31] or "Report")
    columns = list(rows_df.columns)
    merge_end = max(len(columns), 2)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_end)
    worksheet.cell(row=1, column=1, value=title)
    worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=15, color="1F1F1F")
    worksheet.cell(row=1, column=1).fill = openpyxl.styles.PatternFill("solid", fgColor="DCE6F1")

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_end)
    worksheet.cell(row=2, column=1, value=f"Department: {department}" if department else "Department: All Departments")
    worksheet.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True, size=11, color="2C3E50")

    header_row = 4
    for index, column in enumerate(columns, start=1):
        cell = worksheet.cell(row=header_row, column=index, value=column)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="2F5D7C")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin", color="B0B7C3"),
            right=openpyxl.styles.Side(style="thin", color="B0B7C3"),
            top=openpyxl.styles.Side(style="thin", color="B0B7C3"),
            bottom=openpyxl.styles.Side(style="thin", color="B0B7C3"),
        )

    for row_index, (_, row) in enumerate(rows_df.iterrows(), start=header_row + 1):
        for column_index, column_name in enumerate(columns, start=1):
            value = _clean_cell(row[column_name])
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="right" if column_name in _PRICE_COLUMNS | _QUANTITY_COLUMNS else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style="thin", color="D9D9D9"),
                right=openpyxl.styles.Side(style="thin", color="D9D9D9"),
                top=openpyxl.styles.Side(style="thin", color="D9D9D9"),
                bottom=openpyxl.styles.Side(style="thin", color="D9D9D9"),
            )
            if row_index % 2 == 1:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="F7F9FC")
            if column_name in _PRICE_COLUMNS:
                cell.number_format = '$#,##0.00'
            elif column_name in _QUANTITY_COLUMNS:
                cell.number_format = '#,##0.00'

    worksheet.freeze_panes = "A5"
    _autofit_worksheet(worksheet)


def _style_summary_sheet(worksheet) -> None:
    header_fill = openpyxl.styles.PatternFill("solid", fgColor="2F5D7C")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    _autofit_worksheet(worksheet)


def generate_excel_export(
    main_sheet_name: str,
    rows_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    context_column: str | None = None,
    table_columns: Sequence[str] | None = None,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        workbook = writer.book
        summary_sheet = writer.sheets["Summary"]
        _style_summary_sheet(summary_sheet)
        prepared = _prepare_export_groups(
            rows_df,
            context_column=context_column,
            table_columns=table_columns,
            flat_when_multi_context=False,
        )

        for index, (department, group_df) in enumerate(prepared["groups"]):
            if len(prepared["groups"]) == 1:
                sheet_name = main_sheet_name
            else:
                safe_department = sanitize_filename_part(department or f"group_{index + 1}")
                sheet_name = f"{index + 1:02d}_{safe_department}"[:31]
            _write_report_sheet(workbook, sheet_name, main_sheet_name, department, group_df)
    output.seek(0)
    return output.getvalue()


def _format_pdf_value(column: str, value) -> str:
    cleaned = _clean_cell(value)
    if cleaned == "":
        return ""
    if isinstance(cleaned, bool):
        return "Yes" if cleaned else "No"
    if isinstance(cleaned, (int, float)):
        if column in _PRICE_COLUMNS or column.endswith(" Price"):
            return f"${float(cleaned):,.2f}"
        if float(cleaned).is_integer():
            return str(int(float(cleaned)))
        return f"{float(cleaned):,.2f}"
    return str(cleaned)


def _cell_paragraph(text: str, style) -> Paragraph:
    return Paragraph(escape(text).replace("\n", "<br/>"), style)


def _split_dataframe_columns(rows_df: pd.DataFrame, repeat_columns: Sequence[str], max_columns: int) -> List[pd.DataFrame]:
    if rows_df.empty or len(rows_df.columns) <= max_columns:
        return [rows_df]

    fixed_columns = [column for column in repeat_columns if column in rows_df.columns]
    variable_columns = [column for column in rows_df.columns if column not in fixed_columns]
    room_for_variable = max(1, max_columns - len(fixed_columns))
    groups = []
    for index in range(0, len(variable_columns), room_for_variable):
        groups.append(rows_df[fixed_columns + variable_columns[index:index + room_for_variable]].copy())
    return groups or [rows_df]


def _column_weights(columns: Iterable[str]) -> List[float]:
    weights = []
    for column in columns:
        if column.endswith(" Match"):
            weights.append(2.1)
        elif column.endswith(" Price"):
            weights.append(1.15)
        else:
            weights.append(_WIDE_COLUMN_HINTS.get(column, 1.0))
    return weights


def _build_table(rows_df: pd.DataFrame, mode: str, available_width: float):
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        "ExportHeader",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.whitesmoke if mode == "download" else colors.white,
        alignment=1,
    )
    body_style = ParagraphStyle(
        "ExportBody",
        parent=styles["BodyText"],
        fontSize=7,
        leading=9,
        textColor=colors.black,
    )
    numeric_style = ParagraphStyle(
        "ExportBodyNumeric",
        parent=body_style,
        alignment=2,
    )

    headers = list(rows_df.columns)
    table_rows = [[_cell_paragraph(str(column), header_style) for column in headers]]
    for _, row in rows_df.iterrows():
        table_rows.append(
            [
                _cell_paragraph(
                    _format_pdf_value(column, row[column]),
                    numeric_style if column in _PRICE_COLUMNS | _QUANTITY_COLUMNS else body_style,
                )
                for column in headers
            ]
        )

    weights = _column_weights(headers)
    weight_total = sum(weights) or 1.0
    col_widths = [available_width * weight / weight_total for weight in weights]
    table = Table(table_rows, colWidths=col_widths, repeatRows=1, hAlign="LEFT")

    if mode == "print":
        header_background = colors.HexColor("#2b2b2b")
        row_backgrounds = [colors.white, colors.HexColor("#f1f1f1")]
    else:
        header_background = colors.HexColor("#1f77b4")
        row_backgrounds = [colors.white, colors.HexColor("#eef3f8")]

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), header_background),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), row_backgrounds),
            ]
        )
    )
    for index, header in enumerate(headers):
        if header in _PRICE_COLUMNS | _QUANTITY_COLUMNS:
            table.setStyle(TableStyle([("ALIGN", (index, 1), (index, -1), "RIGHT")]))
    return table


def generate_table_pdf(
    title: str,
    project_meta: Dict[str, object],
    rows_df: pd.DataFrame,
    mode: str,
    group_by: str | None = None,
    repeat_columns: Sequence[str] | None = None,
    context_column: str | None = None,
    table_columns: Sequence[str] | None = None,
) -> bytes:
    if not EXPORT_PDF_AVAILABLE:
        raise RuntimeError("PDF export not available")

    output = BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        rightMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        textColor=colors.HexColor("#1f1f1f"),
        alignment=1,
        spaceAfter=8,
    )
    meta_style = ParagraphStyle(
        "ExportMeta",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#333333"),
        alignment=1,
    )
    section_style = ParagraphStyle(
        "ExportSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=13,
        textColor=colors.HexColor("#222222"),
        spaceBefore=4,
        spaceAfter=6,
    )

    elements = [Paragraph(escape(title), title_style)]

    meta_lines = []
    for label, value in project_meta.items():
        rendered = _normalize_filter_value(value)
        if rendered and rendered not in {"All", "None"}:
            meta_lines.append(f"{label}: {rendered}")
    if meta_lines:
        for line in meta_lines:
            elements.append(Paragraph(escape(line), meta_style))
    elements.append(Spacer(1, 0.16 * inch))

    repeat_columns = [column for column in (repeat_columns or []) if not context_column or column != context_column]
    max_columns = 9 if mode == "download" else 7
    available_width = page_size[0] - doc.leftMargin - doc.rightMargin

    prepared = _prepare_export_groups(
        rows_df,
        context_column=context_column,
        table_columns=table_columns,
        flat_when_multi_context=False,
    )
    single_context = prepared["context_value"]
    if single_context:
        elements.append(Paragraph(f"Department: {escape(single_context)}", section_style))
        elements.append(Spacer(1, 0.08 * inch))

    if group_by and len(prepared["groups"]) > 1:
        for group_index, (group_name, group_df) in enumerate(prepared["groups"]):
            if group_index > 0:
                elements.append(PageBreak())
            elements.append(Paragraph(f"{group_by}: {escape(str(group_name))}", section_style))
            column_sets = _split_dataframe_columns(group_df, repeat_columns, max_columns)
            for column_index, table_df in enumerate(column_sets):
                if column_index > 0:
                    elements.append(Spacer(1, 0.1 * inch))
                    elements.append(
                        Paragraph(
                            f"{group_by}: {escape(str(group_name))} (column set {column_index + 1} of {len(column_sets)})",
                            meta_style,
                        )
                    )
                elements.append(_build_table(table_df, mode, available_width))
                elements.append(Spacer(1, 0.14 * inch))
    else:
        table_df = prepared["groups"][0][1]
        column_sets = _split_dataframe_columns(table_df, repeat_columns, max_columns)
        for index, table_df in enumerate(column_sets):
            if index > 0:
                elements.append(PageBreak())
                elements.append(Paragraph(escape(title), title_style))
                for line in meta_lines:
                    elements.append(Paragraph(escape(line), meta_style))
                elements.append(Spacer(1, 0.16 * inch))
            if len(column_sets) > 1:
                elements.append(Paragraph(f"Column Set {index + 1} of {len(column_sets)}", section_style))
            elements.append(_build_table(table_df, mode, available_width))
            elements.append(Spacer(1, 0.14 * inch))

    doc.build(elements)
    output.seek(0)
    return output.getvalue()
